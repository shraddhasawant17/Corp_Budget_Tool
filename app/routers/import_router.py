import os
import json
import uuid
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.database import get_db
from app import models

router    = APIRouter()
templates = Jinja2Templates(directory="app/templates")

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── Master Data (from Excel analysis) ──────────────────
BUSINESS_COST_MAP = {
    "Operation Technology" : 3709,
    "RISK Technology"      : 3708,
    "Technology infra"     : 3701,
    "Technology operations": 3717,
    "Technology others"    : 3722,
    "Technology shared DWH": 3716,
    "Treasury"             : 3714,
}

VALID_OLD_NEW = ["Old", "Old But Incremental", "New"]

VALID_EXPENSE_TYPES = [
    "Hardware", "Managed Services", "Subscription renewal",
    "T&M Services", "Professional Services",
    "Cost of being compliant", "Miscellaneous"
]

# ── Helper: get user from cookie ───────────────────────
async def get_user_from_cookie(request: Request, db: Session):
    from app.auth import decode_token
    from jose import JWTError
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = decode_token(token)
        email   = payload.get("sub")
        return db.query(models.User).filter(models.User.email == email).first()
    except JWTError:
        return None


# ── GET /import/ → Upload Page ─────────────────────────
@router.get("/", response_class=HTMLResponse)
async def import_page(request: Request, db: Session = Depends(get_db)):
    user = await get_user_from_cookie(request, db)
    if not user:
        return RedirectResponse(url="/auth/login")

    # Upload history
    history = db.query(models.UploadedFile).filter(
        models.UploadedFile.uploaded_by == user.id
    ).order_by(models.UploadedFile.upload_at.desc()).limit(10).all()

    return templates.TemplateResponse(request, "import_export/upload.html", {
        "user"       : user,
        "active_page": "import",
        "history"    : history,
        "error"      : None,
    })


# ── POST /import/upload → Parse + Show Summary ─────────
@router.post("/upload")
async def import_upload(
    request : Request,
    db      : Session   = Depends(get_db),
    file    : UploadFile = File(...),
):
    user = await get_user_from_cookie(request, db)
    if not user:
        return RedirectResponse(url="/auth/login")

    # ── Validate file type ──
    if not file.filename.endswith(('.xlsx', '.xls')):
        return templates.TemplateResponse(request, "import_export/upload.html", {
            "user": user, "active_page": "import",
            "history": [], "error": "Only .xlsx or .xls files allowed."
        })

    # ── Save file to disk ──
    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    stored_name  = f"{user.id}{timestamp}{uuid.uuid4().hex[:6]}_{file.filename}"
    file_path    = os.path.join(UPLOAD_DIR, stored_name)
    contents     = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    # ── Parse Excel ──
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Try Sheet1 first, then first sheet
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

        # Find header row (row with "Current year budget key")
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            if row and any(str(v).strip().lower().startswith("current year budget key") for v in row if v):
                header_row_idx = i
                break

        if not header_row_idx:
            # Try Input_Data sheet
            if 'Input_Data' in wb.sheetnames:
                ws = wb['Input_Data']
                header_row_idx = 1
            else:
                raise ValueError("Could not find header row in Excel. Use the provided template.")

        # Get headers
        headers = []
        for row in ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True):
            headers = [str(v).strip() if v else "" for v in row]

        # Column index map
        col_map = {
            "budget_key"    : next((i for i, h in enumerate(headers) if "current year budget key" in h.lower()), None),
            "prev_key"      : next((i for i, h in enumerate(headers) if "previous year budget key" in h.lower()), None),
            "old_new"       : next((i for i, h in enumerate(headers) if h.lower() == "old/new"), None),
            "business"      : next((i for i, h in enumerate(headers) if "business name" in h.lower()), None),
            "cost_code"     : next((i for i, h in enumerate(headers) if "cost code" in h.lower()), None),
            "it_head"       : next((i for i, h in enumerate(headers) if h.lower() == "it head"), None),
            "spoc"          : next((i for i, h in enumerate(headers) if "spoc" in h.lower()), None),
            "exp_desc"      : next((i for i, h in enumerate(headers) if "expense description" in h.lower()), None),
            "budget_a"      : next((i for i, h in enumerate(headers) if "final budget amt" in h.lower() or "(a)" in h.lower()), None),
            "projected_b"   : next((i for i, h in enumerate(headers) if "projected consumption" in h.lower()), None),
            "budget_c"      : next((i for i, h in enumerate(headers) if "budget amt (fy 26" in h.lower() or "( c)" in h.lower() or "(c)" in h.lower()), None),
            "reasoning"     : next((i for i, h in enumerate(headers) if "detailed reasoning" in h.lower()), None),
            "description"   : next((i for i, h in enumerate(headers) if h.lower().strip() == "description"), None),
            "expense_type"  : next((i for i, h in enumerate(headers) if "expense sub type" in h.lower()), None),
            "application"   : next((i for i, h in enumerate(headers) if "application" in h.lower()), None),
            "vendor"        : next((i for i, h in enumerate(headers) if "vendor" in h.lower()), None),
            "resource_count": next((i for i, h in enumerate(headers) if "resource count" in h.lower()), None),
        }

        # ── Parse rows ──
        parsed_rows   = []
        valid_rows    = []
        invalid_rows  = []

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
            # Skip empty rows
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            def get(col_name):
                idx = col_map.get(col_name)
                if idx is None or idx >= len(row):
                    return None
                val = row[idx]
                return str(val).strip() if val is not None else None

            def get_num(col_name):
                val = get(col_name)
                if val is None:
                    return None
                try:
                    return float(str(val).replace(",", "").replace("₹", "").strip())
                except:
                    return None

            row_data = {
                "row_num"       : row_num,
                "budget_key"    : get("budget_key"),
                "prev_key"      : get("prev_key"),
                "old_new"       : get("old_new"),
                "business"      : get("business"),
                "cost_code"     : get_num("cost_code"),
                "it_head"       : get("it_head"),
                "spoc"          : get("spoc"),
                "exp_desc"      : get("exp_desc"),
                "budget_a"      : get_num("budget_a"),
                "projected_b"   : get_num("projected_b"),
                "budget_c"      : get_num("budget_c"),
                "reasoning"     : get("reasoning"),
                "description"   : get("description"),
                "expense_type"  : get("expense_type"),
                "application"   : get("application"),
                "vendor"        : get("vendor"),
                "resource_count": get_num("resource_count"),
                "errors"        : [],
                "warnings"      : [],
            }

            # ── VALIDATIONS ──
            errors = []

            if not row_data["budget_key"]:
                errors.append("Budget Key missing")

            if not row_data["old_new"]:
                errors.append("Old/New missing")
            elif row_data["old_new"] not in VALID_OLD_NEW:
                errors.append(f"Old/New invalid: '{row_data['old_new']}' — allowed: {', '.join(VALID_OLD_NEW)}")

            if not row_data["business"]:
                errors.append("Business Name missing")
            elif row_data["business"] not in BUSINESS_COST_MAP:
                errors.append(f"Business Name '{row_data['business']}' not in master list")

            if row_data["business"] in BUSINESS_COST_MAP:
                expected_cost = BUSINESS_COST_MAP[row_data["business"]]
                if row_data["cost_code"] and int(row_data["cost_code"]) != expected_cost:
                    errors.append(f"Cost Code mismatch: {row_data['cost_code']} — expected {expected_cost} for {row_data['business']}")

            if not row_data["expense_type"]:
                errors.append("Expense Sub Type missing")
            elif row_data["expense_type"] not in VALID_EXPENSE_TYPES:
                errors.append(f"Expense Type '{row_data['expense_type']}' invalid — allowed: {', '.join(VALID_EXPENSE_TYPES)}")

            if row_data["budget_a"] is None:
                errors.append("Final Budget FY 25-26 (A) missing or invalid")
            elif row_data["budget_a"] < 0:
                errors.append("Budget A cannot be negative")

            if row_data["budget_c"] is None:
                errors.append("Budget Amt FY 26-27 (C) missing or invalid")
            elif row_data["budget_c"] < 0:
                errors.append("Budget C cannot be negative")

            # Duplicate key check in DB
            if row_data["budget_key"]:
                exists = db.query(models.BudgetLine).filter(
                    models.BudgetLine.budget_key_current == row_data["budget_key"]
                ).first()
                if exists:
                    errors.append(f"Budget Key '{row_data['budget_key']}' already exists in DB")

            # Warnings (not blocking)
            if not row_data["exp_desc"]:
                row_data["warnings"].append("Expense Description empty")
            if not row_data["reasoning"]:
                row_data["warnings"].append("Detailed Reasoning empty")
            if row_data["projected_b"] is None:
                row_data["warnings"].append("Projected Consumption (B) empty — can be filled later")

            row_data["errors"] = errors
            row_data["valid"]  = len(errors) == 0
            parsed_rows.append(row_data)

            if row_data["valid"]:
                valid_rows.append(row_data)
            else:
                invalid_rows.append(row_data)

        wb.close()

    except Exception as e:
        os.remove(file_path)
        return templates.TemplateResponse(request, "import_export/upload.html", {
            "user": user, "active_page": "import",
            "history": [], "error": f"Excel parse error: {str(e)}"
        })

    # ── Save file record in DB ──
    file_record = models.UploadedFile(
        original_name = file.filename,
        stored_name   = stored_name,
        uploaded_by   = user.id,
        rows_total    = len(parsed_rows),
        rows_imported = 0,
        status        = "pending",
    )
    db.add(file_record)
    db.commit()
    db.refresh(file_record)

    # Pass data to summary page via session-like approach (store in temp JSON)
    summary_data = {
        "file_id"    : file_record.id,
        "file_name"  : file.filename,
        "stored_name": stored_name,
        "total"      : len(parsed_rows),
        "valid_count": len(valid_rows),
        "invalid_count": len(invalid_rows),
        "rows"       : parsed_rows,
    }

    # Store summary in a temp file
    temp_path = os.path.join(UPLOAD_DIR, f"summary_{file_record.id}.json")
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(summary_data, f, ensure_ascii=False, default=str)

    return templates.TemplateResponse(request, "import_export/summary.html", {
        "user"         : user,
        "active_page"  : "import",
        "summary"      : summary_data,
        "file_id"      : file_record.id,
    })


# ── POST /import/confirm → Save Valid Rows to DB ───────
@router.post("/confirm")
async def import_confirm(
    request : Request,
    db      : Session = Depends(get_db),
    file_id : int     = Form(...),
):
    user = await get_user_from_cookie(request, db)
    if not user:
        return RedirectResponse(url="/auth/login")

    # Load summary JSON
    temp_path = os.path.join(UPLOAD_DIR, f"summary_{file_id}.json")
    if not os.path.exists(temp_path):
        return RedirectResponse(url="/import/?error=Session expired. Please upload again.")

    with open(temp_path, "r", encoding="utf-8") as f:
        summary = json.load(f)

    valid_rows    = [r for r in summary["rows"] if r["valid"]]
    imported      = 0
    expense_map   = {e.value: e for e in models.ExpenseSubType}
    old_new_map   = {o.value: o for o in models.OldNew}

    for row in valid_rows:
        # Calculate diffs
        a = row["budget_a"] or 0
        b = row["projected_b"]
        c = row["budget_c"] or 0
        diff_ab = (a - b) if b is not None else None
        diff_cb = (c - b) if b is not None else None
        diff_ca = c - a

        # Cost code from master map
        cost_code = BUSINESS_COST_MAP.get(row["business"], row["cost_code"])

        entry = models.BudgetLine(
            budget_key_current   = row["budget_key"],
            budget_key_previous  = row["prev_key"],
            old_new              = old_new_map.get(row["old_new"], models.OldNew.OLD),
            business_name        = row["business"],
            cost_code            = int(cost_code) if cost_code else 0,
            submitted_by         = user.id,
            it_head_name         = row["it_head"] or user.full_name,
            spoc_name            = row["spoc"],
            expense_sub_type     = expense_map.get(row["expense_type"], models.ExpenseSubType.MISCELLANEOUS),
            expense_description  = row["exp_desc"],
            description          = row["description"],
            application_platform = row["application"],
            vendor_name          = row["vendor"],
            resource_count       = int(row["resource_count"]) if row["resource_count"] else None,
            budget_amt_current_fy= a,
            projected_consumption= b,
            budget_amt_next_fy   = c,
            diff_a_minus_b       = diff_ab,
            diff_c_minus_b       = diff_cb,
            diff_c_minus_a       = diff_ca,
            detailed_reasoning   = row["reasoning"],
            import_file_id       = file_id,
            status               = models.BudgetStatus.DRAFT,
        )
        db.add(entry)
        imported += 1

    # Update file record
    file_rec = db.query(models.UploadedFile).filter(models.UploadedFile.id == file_id).first()
    if file_rec:
        file_rec.rows_imported = imported
        file_rec.status        = "processed"

    db.commit()

    # Cleanup temp file
    try:
        os.remove(temp_path)
    except:
        pass

    return RedirectResponse(url=f"/budget/?imported={imported}", status_code=302)


# ── GET /import/history ────────────────────────────────
@router.get("/history", response_class=HTMLResponse)
async def import_history(request: Request, db: Session = Depends(get_db)):
    user = await get_user_from_cookie(request, db)
    if not user:
        return RedirectResponse(url="/auth/login")

    if user.role == models.UserRole.IT_HEAD:
        history = db.query(models.UploadedFile).filter(
            models.UploadedFile.uploaded_by == user.id
        ).order_by(models.UploadedFile.upload_at.desc()).all()
    else:
        history = db.query(models.UploadedFile).order_by(
            models.UploadedFile.upload_at.desc()
        ).all()

    return templates.TemplateResponse(request, "import_export/upload.html", {
        "user"       : user,
        "active_page": "import",
        "history"    : history,
        "error"      : None,
    })